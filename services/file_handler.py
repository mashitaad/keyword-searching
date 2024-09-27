import os
import pandas as pd
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def format_filename(filename):
    """
    Replaces the character '/' with '-' in the file name.
    
    Args:
        filename (str): The original filename.
        
    Returns:
        str: The formatted filename.
    """
    return filename.replace('/', '-')

def export_results(results, keyword):
    """
    Saves the search results to an Excel file in the 'output/' folder and applies styling.
    
    Args:
        results (list): The list of search results.
        keyword (str): The keyword that was searched.
    """
    df = pd.DataFrame(results, columns=['Filename', 'Keyword'])
    
    safe_keyword = format_filename(keyword)
    
    output_folder = "output/"
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    output_file = os.path.join(output_folder, f"{safe_keyword}.xlsx")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        
        format_sheet(writer)
    
    logging.info(f"Results saved to '{output_file}'")

def format_sheet(writer):
    """
    Applies custom styling to the Excel file, such as bold headers, 
    adjusted column widths, alternating row colors, and borders.
    
    Args:
        writer (pd.ExcelWriter): The pandas Excel writer object.
    """
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    worksheet.column_dimensions['A'].width = 40  # Filename
    worksheet.column_dimensions['B'].width = 60  # Keyword
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    fill_odd = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")
    fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if row[0].row % 2 == 0:
            fill = fill_even
        else:
            fill = fill_odd
        for cell in row:
            cell.fill = fill
            cell.border = thin_border
